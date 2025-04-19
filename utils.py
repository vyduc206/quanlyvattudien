import base64
import os
import re
import pandas as pd
from io import BytesIO
from datetime import datetime
from flask import current_app
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Cấu hình môi trường cho OpenCV
def configure_opencv_environment():
    """Hàm cấu hình môi trường cho OpenCV hoạt động trên các môi trường server"""
    os.environ['OPENCV_IO_ENABLE_JASPER'] = 'true'
    os.environ['OPENCV_IO_MAX_IMAGE_PIXELS'] = str(pow(2, 31))

# Gọi hàm cấu hình trước khi import OpenCV
configure_opencv_environment()

# Import các thư viện có thể gây lỗi
try:
    import numpy as np
    import cv2
    from pyzbar.pyzbar import decode
    CV_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Some image processing features may be limited: {e}")
    CV_AVAILABLE = False

# Hàm tạo mã phiếu tự động
def generate_reference_number(prefix, model):
    """
    Tạo mã tham chiếu cho phiếu nhập/xuất
    """
    today = datetime.now()
    date_part = today.strftime('%Y%m%d')
    
    # Tìm mã phiếu cuối cùng trong ngày
    latest_doc = model.query.filter(
        model.ma_phieu.like(f'{prefix}{date_part}%')
    ).order_by(model.id.desc()).first()
    
    if latest_doc:
        # Lấy số cuối cùng và tăng lên 1
        last_num = int(latest_doc.ma_phieu[-4:])
        new_num = last_num + 1
    else:
        new_num = 1
        
    return f"{prefix}{date_part}{new_num:04d}"

# Hàm xử lý ảnh từ dữ liệu base64
def process_base64_image(base64_image, max_width=800, max_height=600):
    """
    Xử lý ảnh từ dữ liệu base64 và điều chỉnh kích thước
    """
    if not CV_AVAILABLE:
        # Nếu OpenCV không khả dụng, trả về ảnh gốc
        return base64_image
        
    try:
        if base64_image.startswith('data:image'):
            # Trích xuất phần dữ liệu
            base64_data = base64_image.split(',')[1]
        else:
            base64_data = base64_image
            
        # Giải mã base64
        image_data = base64.b64decode(base64_data)
        
        # Chuyển đổi thành mảng numpy
        nparr = np.frombuffer(image_data, np.uint8)
        
        # Đọc ảnh bằng OpenCV
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return base64_image
            
        # Điều chỉnh kích thước nếu cần
        h, w = img.shape[:2]
        if h > max_height or w > max_width:
            # Tính toán tỷ lệ để giữ nguyên tỷ lệ khung hình
            ratio = min(max_width/w, max_height/h)
            new_w = int(w * ratio)
            new_h = int(h * ratio)
            img = cv2.resize(img, (new_w, new_h))
        
        # Chuyển lại thành dữ liệu base64
        _, buffer = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 85])
        jpg_as_text = base64.b64encode(buffer).decode('utf-8')
        
        return f"data:image/jpeg;base64,{jpg_as_text}"
    except Exception as e:
        print(f"Error processing image: {e}")
        return base64_image

# Hàm quét mã vạch từ ảnh base64
def scan_barcode_from_base64(base64_image):
    """
    Quét mã vạch từ ảnh base64
    """
    if not CV_AVAILABLE:
        # Nếu không có thư viện cần thiết, trả về danh sách rỗng
        print("Warning: CV libraries not available for barcode scanning")
        return []
        
    try:
        if base64_image.startswith('data:image'):
            # Trích xuất phần dữ liệu
            base64_data = base64_image.split(',')[1]
        else:
            base64_data = base64_image
            
        # Giải mã base64
        image_data = base64.b64decode(base64_data)
        
        # Chuyển đổi thành mảng numpy
        nparr = np.frombuffer(image_data, np.uint8)
        
        # Đọc ảnh bằng OpenCV
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            print("Warning: Could not decode image for barcode scanning")
            return []
            
        # Chuyển sang ảnh xám
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Quét mã vạch
        barcodes = decode(gray)
        
        results = []
        for barcode in barcodes:
            barcode_data = barcode.data.decode("utf-8")
            barcode_type = barcode.type
            results.append({
                'data': barcode_data,
                'type': barcode_type
            })
        
        return results
    except Exception as e:
        print(f"Error scanning barcode: {e}")
        return []

# Hàm xuất dữ liệu ra Excel
def export_to_excel(data, headers, filename, sheet_name='Sheet1'):
    """
    Xuất dữ liệu ra file Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Thêm tiêu đề
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Thêm dữ liệu
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Điều chỉnh độ rộng cột
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    # Lưu file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Hàm xuất dữ liệu ra PDF
def export_to_pdf(data, headers, title, filename):
    """
    Xuất dữ liệu ra file PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    # Thiết lập style
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Vietnamese', fontName='Helvetica', fontSize=12, leading=12))
    
    # Thêm tiêu đề
    elements.append(Paragraph(title, styles['Heading1']))
    elements.append(Spacer(1, 12))
    
    # Thêm ngày xuất báo cáo
    date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Paragraph(f"Thời gian xuất báo cáo: {date_str}", styles['Vietnamese']))
    elements.append(Spacer(1, 12))
    
    # Tạo bảng
    table_data = [headers]
    for row in data:
        table_data.append(row)
    
    table = Table(table_data, repeatRows=1)
    
    # Thiết lập style cho bảng
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Tạo PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

# Hàm định dạng số tiền
def format_currency(amount):
    """
    Định dạng số tiền theo tiêu chuẩn Việt Nam
    """
    return "{:,.0f}".format(amount).replace(',', '.')

# Hàm tạo slug
def create_slug(text):
    """
    Tạo slug từ văn bản tiếng Việt
    """
    # Chuyển về chữ thường
    text = text.lower()
    # Thay thế các ký tự tiếng Việt
    text = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', text)
    text = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', text)
    text = re.sub(r'[ìíịỉĩ]', 'i', text)
    text = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', text)
    text = re.sub(r'[ùúụủũưừứựửữ]', 'u', text)
    text = re.sub(r'[ỳýỵỷỹ]', 'y', text)
    text = re.sub(r'[đ]', 'd', text)
    # Thay thế các ký tự đặc biệt bằng dấu gạch ngang
    text = re.sub(r'[^a-z0-9]+', '-', text)
    # Loại bỏ các dấu gạch ngang thừa
    text = re.sub(r'^-+|-+$', '', text)
    return text
